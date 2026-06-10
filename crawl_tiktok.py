import asyncio
import json
import os
import requests
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse, parse_qs
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

async def scrape_all_comments(video_url):
    all_comments = []
    api_url = None

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()

        async def handle_response(response):
            nonlocal api_url
            url = response.url
            if "comment/list" in url:
                if not api_url:
                    print("\n[+] FOUND API\n")
                    api_url = url

        page.on("response", handle_response)
        await page.goto(video_url, wait_until="domcontentloaded")
        
        print("[*] Đang tải trang, vui lòng chờ...")
        await page.wait_for_timeout(8000)

        try:
            await page.locator('[data-e2e="comment-icon"]').click(timeout=10000)
            print("Clicked comment icon")
        except:
            pass

        for _ in range(10):
            await page.mouse.wheel(0, 12000)
            await page.wait_for_timeout(2500)
            if api_url:
                break

        if not api_url:
            print("Cannot capture API URL")
            await browser.close()
            return []

        parsed = urlparse(api_url)
        base_url = parsed.scheme + "://" + parsed.netloc + parsed.path
        params = parse_qs(parsed.query)

        cursor = 0
        comment_index = 1
        retry_count = 0 # Thêm biến đếm số lần giải Captcha

        while True:
            params["cursor"] = [str(cursor)]
            flat_params = {k: v[0] for k, v in params.items()}
            print(f"\nFetching cursor {cursor}")

            try:
                data = await page.evaluate(
                    """
                    async ({url, params}) => {
                        const qs = new URLSearchParams(params)
                        const res = await fetch(url + "?" + qs.toString(), {credentials: "include"})
                        const text = await res.text()
                        try {
                            return JSON.parse(text)
                        } catch (e) {
                            return { "error_flag": true }
                        }
                    }
                    """,
                    {"url": base_url, "params": flat_params}
                )
            except Exception as e:
                print("FETCH ERROR:", e)
                break

            # ĐOẠN ĐÃ ĐƯỢC SỬA: CHỜ USER GIẢI CAPTCHA THAY VÌ THOÁT
            if data.get("error_flag"):
                retry_count += 1
                if retry_count > 3:
                    print("[-] Kẹt Captcha quá nhiều lần, hệ thống tự động dừng để an toàn.")
                    break
                    
                print("\n[!!!] BÁO ĐỘNG: TikTok đòi xác minh CAPTCHA!")
                print("[!!!] BẠN CÓ 20 GIÂY ĐỂ KÉO THANH TRƯỢT TRÊN TRÌNH DUYỆT...")
                await asyncio.sleep(20) # Tạm dừng 20 giây cho bạn thao tác
                print("[*] Đã hết 20 giây, hệ thống thử tiếp tục cào...\n")
                continue # Bỏ qua vòng lặp này và thử gọi lại API (không break nữa)

            # Nếu lấy được data bình thường thì reset bộ đếm
            retry_count = 0 
            comments = data.get("comments", [])
            print(f"Received {len(comments)} comments")

            if not comments:
                print("No comments returned")
                break

            for c in comments:
                user = c.get("user", {})
                timestamp = c.get("create_time")
                created_time = datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

                # Xử lý tải hình ảnh
                image_urls = []
                for img in c.get("image_list") or []:
                    try:
                        image_urls.append(img["origin_url"]["url_list"][0])
                    except:
                        pass

                local_image_path = ""
                if image_urls:
                    try:
                        os.makedirs("comment_images", exist_ok=True)
                        image_url = image_urls[0]
                        filename = f"comment_images/{comment_index}.jpg"
                        r = requests.get(image_url)
                        with open(filename, "wb") as f:
                            f.write(r.content)
                        local_image_path = filename
                    except Exception as e:
                        print("IMAGE DOWNLOAD ERROR:", e)

                # Xử lý Sticker và Voice
                sticker_url = None
                sticker = c.get("sticker")
                if sticker:
                    try:
                        sticker_url = sticker["animate_url"]["url_list"][0]
                    except:
                        try:
                            sticker_url = sticker["static_url"]["url_list"][0]
                        except:
                            pass

                voice_url = None
                voice_info = c.get("voice_info")
                if voice_info:
                    try:
                        voice_url = voice_info["play_url"]["url_list"][0]
                    except:
                        pass

                all_comments.append({
                    "id": comment_index,
                    "comment_id": c.get("cid"),
                    "created_time": created_time,
                    "text": c.get("text"),
                    "username": user.get("unique_id"),
                    "nickname": user.get("nickname"),
                    "user_id": user.get("uid"),
                    "likes": c.get("digg_count"),
                    "replies": c.get("reply_comment_total"),
                    "image_file": local_image_path,
                    "sticker_or_gif": sticker_url,
                    "voice_comment": voice_url
                })
                comment_index += 1

            has_more = data.get("has_more")
            if not has_more:
                print("\nNo more comments")
                break

            cursor = data.get("cursor")
            await asyncio.sleep(2)

        await browser.close()
    return all_comments

async def main():
    url = "https://vt.tiktok.com/ZSQB4ceXc/"
    comments = await scrape_all_comments(url)
    print(f"\nTOTAL COMMENTS: {len(comments)}")

    with open("tiktok_comments.json", "w", encoding="utf-8") as f:
        json.dump(comments, f, ensure_ascii=False, indent=4)
    print("\nSaved tiktok_comments.json")

    excel_file = "tiktok_comments.xlsx"
    df = pd.DataFrame(comments)
    df.to_excel(excel_file, index=False)

    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["J"].width = 25

        for row in range(2, len(comments) + 2):
            image_path = ws[f"J{row}"].value
            if image_path and os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 100
                    img.height = 100
                    ws.add_image(img, f"J{row}")
                    ws.row_dimensions[row].height = 80
                except Exception as e:
                    print("EXCEL IMAGE ERROR:", e)
        wb.save(excel_file)
        print("Saved tiktok_comments.xlsx with images!")
    except Exception as e:
        print("Không thể chèn ảnh vào Excel:", e)

if __name__ == "__main__":
    asyncio.run(main())